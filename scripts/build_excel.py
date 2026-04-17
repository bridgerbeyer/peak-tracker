import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
import json, sys

# ── Colors ──────────────────────────────────────────────────────────────────
RED       = "CC2222"
DARK      = "1A1A1A"
GRAY_HDR  = "2C2C2E"
GRAY_LIGHT= "F5F5F5"
GREEN     = "059669"
AMBER     = "D97706"
BLUE_LINK = "2563EB"
WHITE     = "FFFFFF"

# ── Helpers ─────────────────────────────────────────────────────────────────
def hdr_font(size=10, bold=True, color=WHITE):
    return Font(name="Arial", size=size, bold=bold, color=color)

def body_font(size=10, bold=False, color=DARK):
    return Font(name="Arial", size=size, bold=bold, color=color)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def border():
    thin = Side(style="thin", color="E0E0E0")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def currency_fmt():
    return '$#,##0;($#,##0);"-"'

def pct_fmt():
    return '0.0%'

def set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

def write_hdr(ws, row, col, value, bg=GRAY_HDR, font_color=WHITE, bold=True, align="center"):
    c = ws.cell(row=row, column=col, value=value)
    c.font = hdr_font(bold=bold, color=font_color)
    c.fill = fill(bg)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    c.border = border()
    return c

def write_cell(ws, row, col, value, bold=False, color=DARK, align="left", number_format=None, bg=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = body_font(bold=bold, color=color)
    c.alignment = Alignment(horizontal=align, vertical="center")
    c.border = border()
    if number_format:
        c.number_format = number_format
    if bg:
        c.fill = fill(bg)
    return c

# ── Load real data from route ────────────────────────────────────────────────
if len(sys.argv) < 3:
    print("Usage: build_excel.py <input.json> <output.xlsx>", file=sys.stderr)
    sys.exit(1)

with open(sys.argv[1], encoding="utf-8") as f:
    rows = json.load(f)

out_path = sys.argv[2]

wb = Workbook()

# ════════════════════════════════════════════════════════════════════════════
# SHEET 1 — Sales Pipeline
# ════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Sales Pipeline"
ws1.freeze_panes = "A4"
ws1.sheet_view.showGridLines = False
ws1.row_dimensions[1].height = 40
ws1.row_dimensions[2].height = 18
ws1.row_dimensions[3].height = 28

# Title row
ws1.merge_cells("A1:N1")
title = ws1["A1"]
title.value = "PEAK CONDO STORAGE — Sales Pipeline"
title.font = Font(name="Arial", size=14, bold=True, color=WHITE)
title.fill = fill(RED)
title.alignment = Alignment(horizontal="left", vertical="center", indent=1)

# Subtitle row
ws1.merge_cells("A2:N2")
sub = ws1["A2"]
sub.value = f"Construction Knowledge Base Export"
sub.font = Font(name="Arial", size=9, color="888888")
sub.alignment = Alignment(horizontal="left", vertical="center", indent=1)

# Column headers
headers = [
    ("Unit", 8),("Phase", 10),("Status", 14),("Buyer", 16),("Close Date", 12),
    ("List Price", 13),("Comm %", 8),("Commission", 13),
    ("Costs", 13),("Add-on COs", 13),("Included COs", 13),
    ("Final Price", 13),("Net to Seller", 14),("Notes", 30),
]
for i, (hdr, width) in enumerate(headers, 1):
    write_hdr(ws1, 3, i, hdr)
    set_col_width(ws1, i, width)

# Status colors
status_bg = {"Sold": "D1FAE5", "Under Contract": "FEF3C7", "Available": "F3F4F6"}
status_fc = {"Sold": "065F46",  "Under Contract": "92400E", "Available": "374151"}

# Data rows
for r, row in enumerate(rows, 4):
    bg = status_bg.get(row["status"], "FFFFFF")
    fc = status_fc.get(row["status"], DARK)
    stripe = "FAFAFA" if r % 2 == 0 else WHITE

    write_cell(ws1, r, 1, row["unit"], bold=True)
    write_cell(ws1, r, 2, row["phase"])
    c_status = write_cell(ws1, r, 3, row["status"], bold=True, color=fc, bg=bg, align="center")
    write_cell(ws1, r, 4, row["buyer"] or "—")
    write_cell(ws1, r, 5, row["closeDate"] or "—", align="center")
    write_cell(ws1, r, 6, row["price"] if row["price"] else None, align="right", number_format=currency_fmt())
    write_cell(ws1, r, 7, (row["commissionPct"] or 0) / 100, align="right", number_format=pct_fmt())
    write_cell(ws1, r, 8, row["commission"] if row["commission"] else None, align="right", number_format=currency_fmt(), color="CC2222" if row["commission"] else DARK)
    write_cell(ws1, r, 9, row["costs"] if row["costs"] else None, align="right", number_format=currency_fmt(), color="CC2222" if row["costs"] else DARK)
    write_cell(ws1, r, 10, row["addOnCOs"] if row["addOnCOs"] else None, align="right", number_format=currency_fmt(), color="CC2222" if row["addOnCOs"] else DARK)
    write_cell(ws1, r, 11, row["includedCOs"] if row["includedCOs"] else None, align="right", number_format=currency_fmt())
    write_cell(ws1, r, 12, row["finalPrice"] if row["finalPrice"] else None, bold=True, align="right", number_format=currency_fmt())
    net_color = GREEN if row["net"] > 0 else "CC2222"
    write_cell(ws1, r, 13, row["net"] if row["net"] else None, bold=True, align="right", number_format=currency_fmt(), color=net_color if row["net"] else DARK)
    notes = " | ".join(filter(None, [row["costBreakdown"], row["coBreakdown"]]))
    write_cell(ws1, r, 14, notes or "—", color="666666")

# Summary section
last_data_row = 3 + len(rows)
summary_row = last_data_row + 2

ws1.merge_cells(f"A{summary_row}:E{summary_row}")
sh = ws1.cell(row=summary_row, column=1, value="SUMMARY")
sh.font = Font(name="Arial", size=10, bold=True, color=WHITE)
sh.fill = fill(GRAY_HDR)
sh.alignment = Alignment(horizontal="left", vertical="center", indent=1)

summary_items = [
    ("Total Units", f'=COUNTA(A4:A{last_data_row})'),
    ("Sold", f'=COUNTIF(C4:C{last_data_row},"Sold")'),
    ("Under Contract", f'=COUNTIF(C4:C{last_data_row},"Under Contract")'),
    ("Available", f'=COUNTIF(C4:C{last_data_row},"Available")'),
    ("Total Pipeline Value", f'=SUMIF(C4:C{last_data_row},"Under Contract",L4:L{last_data_row})'),
    ("Total Closed Value", f'=SUMIF(C4:C{last_data_row},"Sold",L4:L{last_data_row})'),
    ("Avg Net per Sold Unit", f'=IFERROR(SUMIF(C4:C{last_data_row},"Sold",M4:M{last_data_row})/COUNTIF(C4:C{last_data_row},"Sold"),"-")'),
    ("Total Net (All Units)", f'=SUM(M4:M{last_data_row})'),
]

for i, (label, formula) in enumerate(summary_items):
    sr = summary_row + 1 + i
    lc = ws1.cell(row=sr, column=1, value=label)
    lc.font = body_font(bold=True)
    lc.alignment = Alignment(horizontal="left", vertical="center")
    lc.border = border()
    lc.fill = fill(GRAY_LIGHT)
    ws1.merge_cells(f"A{sr}:E{sr}")

    vc = ws1.cell(row=sr, column=6, value=formula)
    vc.font = body_font(color=DARK)
    vc.alignment = Alignment(horizontal="right", vertical="center")
    vc.border = border()
    if "Value" in label or "Net" in label or "Avg" in label:
        vc.number_format = currency_fmt()
    ws1.merge_cells(f"F{sr}:H{sr}")

# ════════════════════════════════════════════════════════════════════════════
# SHEET 2 — By Phase
# ════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("By Phase")
ws2.sheet_view.showGridLines = False

ws2.merge_cells("A1:F1")
t2 = ws2["A1"]
t2.value = "Phase Breakdown"
t2.font = Font(name="Arial", size=13, bold=True, color=WHITE)
t2.fill = fill(RED)
t2.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws2.row_dimensions[1].height = 36

phases = sorted(set(r["phase"] for r in rows))
start_row = 3
for phase in phases:
    phase_rows = [r for r in rows if r["phase"] == phase]
    # Phase header
    ws2.merge_cells(f"A{start_row}:F{start_row}")
    ph = ws2.cell(row=start_row, column=1, value=phase.upper())
    ph.font = Font(name="Arial", size=10, bold=True, color=WHITE)
    ph.fill = fill(GRAY_HDR)
    ph.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws2.row_dimensions[start_row].height = 22
    start_row += 1

    # Col headers
    for i, h in enumerate(["Unit","Status","Buyer","List Price","Net","Close Date"], 1):
        write_hdr(ws2, start_row, i, h, bg="3A3A3C")
    start_row += 1

    for row in phase_rows:
        bg = status_bg.get(row["status"], "FFFFFF")
        fc = status_fc.get(row["status"], DARK)
        write_cell(ws2, start_row, 1, row["unit"], bold=True)
        write_cell(ws2, start_row, 2, row["status"], color=fc, bg=bg, align="center", bold=True)
        write_cell(ws2, start_row, 3, row["buyer"] or "—")
        write_cell(ws2, start_row, 4, row["finalPrice"] if row["finalPrice"] else None, number_format=currency_fmt(), align="right")
        net_color = GREEN if row["net"] > 0 else "CC2222"
        write_cell(ws2, start_row, 5, row["net"] if row["net"] else None, number_format=currency_fmt(), align="right", color=net_color if row["net"] else DARK, bold=True)
        write_cell(ws2, start_row, 6, row["closeDate"] or "—", align="center")
        start_row += 1

    start_row += 1  # gap between phases

for i, w in enumerate([8,14,16,13,14,12], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

# ════════════════════════════════════════════════════════════════════════════
# SHEET 3 — Cost Detail
# ════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Cost Detail")
ws3.sheet_view.showGridLines = False

ws3.merge_cells("A1:G1")
t3 = ws3["A1"]
t3.value = "Cost & Concession Detail"
t3.font = Font(name="Arial", size=13, bold=True, color=WHITE)
t3.fill = fill(RED)
t3.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws3.row_dimensions[1].height = 36

for i, h in enumerate(["Unit","Phase","Status","Cost/CO Item","Amount","Type","Included in Price?"], 1):
    write_hdr(ws3, 3, i, h)

widths = [8,10,14,28,13,14,16]
for i, w in enumerate(widths, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

r = 4
for row in rows:
    bg = status_bg.get(row["status"], "FFFFFF")
    fc = status_fc.get(row["status"], DARK)

    for item in row.get("costItems", []):
        write_cell(ws3, r, 1, row["unit"], bold=True)
        write_cell(ws3, r, 2, row["phase"])
        write_cell(ws3, r, 3, row["status"], color=fc, bg=bg, align="center")
        write_cell(ws3, r, 4, item["label"])
        write_cell(ws3, r, 5, item["amount"], number_format=currency_fmt(), align="right", color="CC2222")
        write_cell(ws3, r, 6, "Cost")
        write_cell(ws3, r, 7, "N/A", align="center", color=DARK)
        r += 1

    for item in row.get("coItems", []):
        included = "Yes" if item.get("included") else "No"
        write_cell(ws3, r, 1, row["unit"], bold=True)
        write_cell(ws3, r, 2, row["phase"])
        write_cell(ws3, r, 3, row["status"], color=fc, bg=bg, align="center")
        write_cell(ws3, r, 4, item["label"])
        write_cell(ws3, r, 5, item["amount"], number_format=currency_fmt(), align="right", color="CC2222")
        write_cell(ws3, r, 6, "Change Order")
        write_cell(ws3, r, 7, included, align="center",
                   color=GREEN if included == "Yes" else "CC2222")
        r += 1

wb.save(out_path)
print("Excel built successfully")
